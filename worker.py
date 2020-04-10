import urllib.request
import io
import openpyxl
import datetime
import os
import redis

from itertools import chain
from collections import defaultdict
from jinja2 import Environment, FileSystemLoader
from time import sleep
from hashlib import md5

try:
    rdb = redis.from_url(os.environ.get('REDIS_URL'))
except:
    rdb = redis.Redis()

def is_floor_shift():
    shifts = set(['... snip ...'])
    return lambda shift: shift in shifts
is_floor_shift = is_floor_shift()

def is_mainrx_shift():
    shifts = set(['... snip ...'])
    return lambda shift: shift in shifts
is_mainrx_shift = is_mainrx_shift()

def hours():
    h = defaultdict(str)
    """ ... snip ... """
    return lambda shift: h[shift]
hours = hours()

def area_covered():
    a = defaultdict(str)
    """ ... snip ... """
    return lambda shift: a[shift]
area_covered = area_covered()

def phone():
    p = defaultdict(str)
    """ ... snip """
    return lambda shift: p[shift]
phone = phone()

def comments():
    c = defaultdict(list)
    """ ... snip ... """
    return lambda shift: c[shift]
comments = comments()

def rounds():
    r = defaultdict(list)
    """ ... snip ... """
    return lambda shift: r[shift]
rounds = rounds()

def normalize_name(name):
    """
    cleans names, removes stuff like - lead
    str -> str
    """
    name = name.strip()
    """ ... snip ... """
    return name

def normalize_shift(shift):
    """
    takes a shift string, if there are slashes it creates two shifts
    str -> [str]
    """
    shifts = shift.split('/')
    res = []
    for s in shifts:
        s = s.strip()
        if not s: continue
        res.append(s)
    return res

def combine_split_shifts(shifts):
    sick = list((shift, name, sick_call) for (shift, name, sick_call) in shifts if sick_call)
    non_sick = list((shift, name, sick_call) for (shift, name, sick_call) in shifts if not sick_call)
    seen = {}
    for (shift, name, _) in non_sick:
        if shift in seen: seen[shift] += ', ' + name
        else: seen[shift] = name
    return list(chain(((shift, name, False) for (shift, name) in seen.items()), sick))

def filter_split(function, iterable):
    """ returns two lists, (if_true, if_false) """
    if_true = []
    if_false = []
    for it in iterable:
        if function(it): if_true.append(it)
        else: if_false.append(it)
    return (if_true, if_false)

class WorkbookCache:
    def __init__(self, url):
        self._url = url
        self._last_hash = None
        self._last_modified_time = None
        self._workbook = None
        self._last_polled_time = None
    def get(self):
        return self._workbook
    def poll(self):
        with urllib.request.urlopen(self._url) as response:
            binary = response.read()
            now = datetime.datetime.now()
            new_hash = md5(binary).hexdigest()
            self._last_polled_time = now
            if new_hash == self._last_hash: 
                return False
            self._last_modified_time = now
            self._last_hash = new_hash
            self._workbook = openpyxl.load_workbook(io.BytesIO(binary), read_only=True)
            return True
    def last_polled(self):
        return self._last_polled_time.strftime('%x %X')
    def last_modified(self):
        return self._last_modified_time.strftime('%x %X')
    def last_hash(self):
        return self._last_hash

class Schedule:
    def __init__(self):
        self._cache = WorkbookCache(self.URL)
    def shifts(self, sheet, column):
        """ yields (shift: str, name: str, sick_call: bool) """
        for row in range(self.ROW_START, self.ROW_END + 1):
            # read cell values
            shift = sheet.cell(row=row, column=column).value
            name = sheet[self.NAME_COLUMN + str(row)].value
            sick_call = sheet.cell(row=row, column=column).fill.fgColor.rgb == 'FFFF0000'
            # skip empty cells
            if not shift or shift == '-': continue
            # clean name
            name = normalize_name(name)
            # some cells have multiple shifts
            for shift in normalize_shift(shift):
                yield (shift, name, sick_call)
    def last_polled(self):
        return self._cache.last_polled()
    def last_modified(self):
        return self._cache.last_modified()
    def assignments(self, date):
        self.workbook = self._cache.get()
        sheet = self.sheet_for_date(date)
        column = self.column_for_date(date)
        shifts = list(self.shifts(sheet, column))
        shifts = combine_split_shifts(shifts)
        return shifts
    def poll(self):
        return self._cache.poll()
    def hash_sum(self):
        return self._cache.last_hash()

class TechSchedule(Schedule):
    URL = '... snip ....'
    NAME_COLUMN = 'B'
    ROW_START = 2
    ROW_END = 40
    def column_for_date(self, date):
        if datetime.date(2020, 2, 1) <= date <= datetime.date(2020, 2, 29):
            return date.day + 1
        if datetime.date(2020, 3, 1) <= date <= datetime.date(2020, 3, 29):
            return date.day + 2
        raise ValueError('date out of range or unsupported')
    def sheet_for_date(self, date):
        if datetime.date(2020, 2, 1) <= date <= datetime.date(2020, 2, 29):
            return self.workbook['Feb 2 2020-Mar1 2020']
        if datetime.date(2020, 3, 1) <= date <= datetime.date(2020, 3, 29):
            return self.workbook['Mar 1 2020-Mar 29 2020']
        raise ValueError('date out of range or unsupported')

class RphSchedule(Schedule):
    URL = '... snip ...'
    NAME_COLUMN = 'A'
    ROW_START = 3
    ROW_END = 43
    def column_for_date(self, date):
        if datetime.date(2020, 2, 2) <= date <= datetime.date(2020, 2, 29):
            return date.day
        if datetime.date(2020, 3, 1) <= date <= datetime.date(2020, 3, 29):
            return date.day + 1
        raise ValueError('date out of range or unsupported')
    def sheet_for_date(self, date):
        if datetime.date(2020, 2, 2) <= date <= datetime.date(2020, 2, 29):
            return self.workbook['Feb 2 - Mar 1 2020']
        if datetime.date(2020, 3, 1) <= date <= datetime.date(2020, 3, 29):
            return self.workbook['Mar 1 - 29 2020 (POSTED) ']
        raise ValueError('date out of range or unsupported')

def render(tech_schedule, rph_schedule, template, date, generated_time):
    tech_shifts = list(tech_schedule.assignments(date))
    rph_shifts = list(rph_schedule.assignments(date))
    sick_calls = list(sorted(_ for _ in chain(tech_shifts, rph_shifts) if _[2]))
    floor_shifts = list(sorted(_ for _ in chain(tech_shifts, rph_shifts) if is_floor_shift(_[0]) and not _[2]))
    mainrx_shifts = list(sorted(_ for _ in rph_shifts if is_mainrx_shift(_[0]) and not _[2]))
    tech_shifts = list(sorted(_ for _ in tech_shifts if not is_floor_shift(_[0]) and not _[2]))

    date_prev = date - datetime.timedelta(days=1)
    date_next = date + datetime.timedelta(days=1)
    date_string = date.strftime('%x %A')

    return template.render(date_prev=date_prev, date_next=date_next, date=date_string,
        floor_shifts=floor_shifts, mainrx_shifts=mainrx_shifts, tech_shifts=tech_shifts,
        hours=hours, phone=phone, area_covered=area_covered, comments=comments, rounds=rounds,
        sick_calls=sick_calls,
        generated_time=generated_time
    )

def date_range(start, end):
    date = start
    while date <= end:
        yield date
        date += datetime.timedelta(days=1)

def main():
    tech_schedule = TechSchedule()
    rph_schedule = RphSchedule()
    template = Environment(loader=FileSystemLoader('templates')).get_template('schedule.html')
    while True:
        print('Polling')
        if any([tech_schedule.poll(), rph_schedule.poll()]):
            print('Updates found')
            gen_time = datetime.datetime.now().strftime('%x %X')
            for date in date_range(datetime.date(2020, 2, 2), datetime.date(2020, 3, 29)):
                html = render(tech_schedule, rph_schedule, template, date, gen_time)
                rdb.set(str(date), html)
                print(f'Generated {date}')
        else:
            print('No updates')
        print('Sleeping')
        sleep(5 * 60)

if __name__ == '__main__':
    main()